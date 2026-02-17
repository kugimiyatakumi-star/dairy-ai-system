# core/excel_io.py
import os
from typing import Dict
import pandas as pd
from openpyxl import load_workbook

def read_sheet_df(xlsx_path: str, sheet: str) -> pd.DataFrame:
    if not os.path.exists(xlsx_path):
        return pd.DataFrame()
    wb = load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet]
    rows = list(ws.values)
    if len(rows) <= 1:
        return pd.DataFrame(columns=list(rows[0]) if rows else [])
    return pd.DataFrame(rows[1:], columns=rows[0])

def append_row(xlsx_path: str, sheet: str, row: Dict):
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(xlsx_path)

    wb = load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        wb.create_sheet(sheet)
    ws = wb[sheet]

    # header (1st row)
    if ws.max_row < 1 or ws.max_column < 1 or ws.cell(1, 1).value is None:
        header = list(row.keys())
        ws.append(header)
    else:
        header = [c.value for c in ws[1]]

    out = [row.get(h, "") for h in header]
    ws.append(out)
    wb.save(xlsx_path)
